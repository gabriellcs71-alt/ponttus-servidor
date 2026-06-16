from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import sqlite3
import io
import calendar
import datetime as _dt
import hashlib
import os
import re
import secrets
import time
import json
from functools import wraps

app = Flask(__name__)
CORS(app)

DB           = os.environ.get('DB_PATH', 'ponto.db')
ADMIN_KEY    = os.environ.get('ADMIN_API_KEY', '')  # Defina esta var para proteger rotas admin
DATABASE_URL = os.environ.get('DATABASE_URL', '')   # Render injeta isto ao linkar um Postgres
USE_PG       = bool(DATABASE_URL)                    # Postgres em produção, SQLite local

# Rate limiting simples em memória: {ip: [timestamps]}
_login_attempts: dict = {}
LOGIN_MAX = 10
LOGIN_WINDOW = 60  # segundos

TOKEN_VALIDADE_DIAS = 90  # token expira após 90 dias sem uso

# Tipos de "abono" que SOMENTE o admin pode lançar (funcionário nunca grava estes).
# '' (vazio) = dia normal de trabalho, preenchido pelo próprio funcionário no app.
ABONO_TIPOS = ('falta', 'atestado', 'folga', 'feriado')

if USE_PG:
    import psycopg
    from psycopg.rows import dict_row
    INTEGRITY_ERRORS = (sqlite3.IntegrityError, psycopg.IntegrityError)
else:
    INTEGRITY_ERRORS = (sqlite3.IntegrityError,)

# Expressões de data/hora por backend — ambas produzem ISO 'YYYY-MM-DD HH:MM:SS' (UTC),
# garantindo que comparações por texto e o formato de saída sejam idênticos nos dois bancos.
if USE_PG:
    NOW_SQL    = "to_char(now() at time zone 'utc','YYYY-MM-DD HH24:MI:SS')"
    CUTOFF_SQL = ("to_char((now() - interval '%d days') at time zone 'utc',"
                  "'YYYY-MM-DD HH24:MI:SS')") % TOKEN_VALIDADE_DIAS
else:
    NOW_SQL    = "datetime('now')"
    CUTOFF_SQL = "datetime('now','-%d days')" % TOKEN_VALIDADE_DIAS

# ── BANCO DE DADOS ─────────────────────────────────────
class _Conn:
    """Adaptador fino: mesma API (.execute(...).fetchone()/fetchall(), .commit(),
    .rollback(), .close()) funcionando tanto em SQLite quanto em Postgres.
    Em Postgres, traduz os placeholders '?' para '%s' automaticamente."""
    def __init__(self):
        if USE_PG:
            self._c = psycopg.connect(DATABASE_URL, row_factory=dict_row)
        else:
            self._c = sqlite3.connect(DB)
            self._c.row_factory = sqlite3.Row

    def execute(self, sql, params=()):
        if USE_PG:
            cur = self._c.cursor()
            cur.execute(sql.replace('?', '%s'), tuple(params) or None)
            return cur
        return self._c.execute(sql, params)

    def commit(self):
        self._c.commit()

    def rollback(self):
        self._c.rollback()

    def close(self):
        self._c.close()

def get_db():
    return _Conn()

def _schema_sql():
    """DDL específico do backend. Datas/timestamps são TEXT (ISO) em ambos."""
    if USE_PG:
        return [
            """CREATE TABLE IF NOT EXISTS funcionarios (
                id SERIAL PRIMARY KEY,
                nome TEXT NOT NULL,
                usuario TEXT UNIQUE NOT NULL,
                senha_hash TEXT NOT NULL,
                matricula TEXT DEFAULT '',
                cargo TEXT DEFAULT '',
                roteirizacao INTEGER DEFAULT 0,
                ativo INTEGER DEFAULT 1,
                criado_em TEXT DEFAULT (to_char(now() at time zone 'utc','YYYY-MM-DD HH24:MI:SS'))
            )""",
            """CREATE TABLE IF NOT EXISTS clientes_geo (
                cod TEXT PRIMARY KEY,
                lat REAL,
                lng REAL,
                precisao TEXT DEFAULT '',
                atualizado_em TEXT DEFAULT ''
            )""",
            """CREATE TABLE IF NOT EXISTS registros (
                id SERIAL PRIMARY KEY,
                funcionario_id INTEGER NOT NULL REFERENCES funcionarios(id),
                data TEXT NOT NULL,
                cidade TEXT,
                entrada TEXT,
                cafe_manha_inicio TEXT,
                cafe_manha_fim TEXT,
                almoco_inicio TEXT,
                almoco_fim TEXT,
                cafe_tarde_inicio TEXT,
                cafe_tarde_fim TEXT,
                jantar_inicio TEXT,
                jantar_fim TEXT,
                saida TEXT,
                observacao TEXT,
                tipo TEXT DEFAULT '',
                enviado_em TEXT DEFAULT (to_char(now() at time zone 'utc','YYYY-MM-DD HH24:MI:SS'))
            )""",
            """CREATE TABLE IF NOT EXISTS tokens (
                token TEXT PRIMARY KEY,
                funcionario_id INTEGER NOT NULL REFERENCES funcionarios(id),
                criado_em TEXT DEFAULT (to_char(now() at time zone 'utc','YYYY-MM-DD HH24:MI:SS')),
                ultimo_uso TEXT DEFAULT (to_char(now() at time zone 'utc','YYYY-MM-DD HH24:MI:SS'))
            )""",
            """CREATE TABLE IF NOT EXISTS checklists (
                id SERIAL PRIMARY KEY,
                funcionario_id INTEGER NOT NULL REFERENCES funcionarios(id),
                data TEXT NOT NULL,
                veiculo TEXT DEFAULT '',
                km TEXT DEFAULT '',
                prox_troca_oleo TEXT DEFAULT '',
                itens TEXT DEFAULT '',
                criado_em TEXT DEFAULT (to_char(now() at time zone 'utc','YYYY-MM-DD HH24:MI:SS'))
            )""",
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_check_func_data
                ON checklists(funcionario_id, data)""",
            """CREATE TABLE IF NOT EXISTS manutencoes (
                id SERIAL PRIMARY KEY,
                funcionario_id INTEGER NOT NULL REFERENCES funcionarios(id),
                data TEXT NOT NULL,
                veiculo TEXT DEFAULT '',
                item TEXT NOT NULL,
                obs TEXT DEFAULT '',
                status TEXT DEFAULT 'pendente',
                criado_em TEXT DEFAULT (to_char(now() at time zone 'utc','YYYY-MM-DD HH24:MI:SS')),
                resolvido_em TEXT DEFAULT ''
            )""",
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_manut_func_data_item
                ON manutencoes(funcionario_id, data, item)""",
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_reg_func_data
                ON registros(funcionario_id, data)""",
        ]
    return [
        """CREATE TABLE IF NOT EXISTS funcionarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            usuario TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            matricula TEXT DEFAULT '',
            cargo TEXT DEFAULT '',
            roteirizacao INTEGER DEFAULT 0,
            ativo INTEGER DEFAULT 1,
            criado_em TEXT DEFAULT (datetime('now'))
        )""",
        """CREATE TABLE IF NOT EXISTS clientes_geo (
            cod TEXT PRIMARY KEY,
            lat REAL,
            lng REAL,
            precisao TEXT DEFAULT '',
            atualizado_em TEXT DEFAULT ''
        )""",
        """CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            funcionario_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            cidade TEXT,
            entrada TEXT,
            cafe_manha_inicio TEXT,
            cafe_manha_fim TEXT,
            almoco_inicio TEXT,
            almoco_fim TEXT,
            cafe_tarde_inicio TEXT,
            cafe_tarde_fim TEXT,
            jantar_inicio TEXT,
            jantar_fim TEXT,
            saida TEXT,
            observacao TEXT,
            tipo TEXT DEFAULT '',
            enviado_em TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        )""",
        """CREATE TABLE IF NOT EXISTS tokens (
            token TEXT PRIMARY KEY,
            funcionario_id INTEGER NOT NULL,
            criado_em TEXT DEFAULT (datetime('now')),
            ultimo_uso TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        )""",
        """CREATE TABLE IF NOT EXISTS checklists (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            funcionario_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            veiculo TEXT DEFAULT '',
            km TEXT DEFAULT '',
            prox_troca_oleo TEXT DEFAULT '',
            itens TEXT DEFAULT '',
            criado_em TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        )""",
        """CREATE UNIQUE INDEX IF NOT EXISTS idx_check_func_data
            ON checklists(funcionario_id, data)""",
        """CREATE TABLE IF NOT EXISTS manutencoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            funcionario_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            veiculo TEXT DEFAULT '',
            item TEXT NOT NULL,
            obs TEXT DEFAULT '',
            status TEXT DEFAULT 'pendente',
            criado_em TEXT DEFAULT (datetime('now')),
            resolvido_em TEXT DEFAULT '',
            FOREIGN KEY (funcionario_id) REFERENCES funcionarios(id)
        )""",
        """CREATE UNIQUE INDEX IF NOT EXISTS idx_manut_func_data_item
            ON manutencoes(funcionario_id, data, item)""",
        """CREATE UNIQUE INDEX IF NOT EXISTS idx_reg_func_data
            ON registros(funcionario_id, data)""",
    ]

def init_db():
    conn = get_db()
    for stmt in _schema_sql():
        conn.execute(stmt)
    conn.commit()  # consolida o schema antes das migrações (evita rollback derrubar as tabelas no PG)

    # Migração (somente SQLite): converte datas antigas DD/MM/YYYY para ISO YYYY-MM-DD
    if not USE_PG:
        conn.execute("""
            UPDATE registros
            SET data = substr(data,7,4) || '-' || substr(data,4,2) || '-' || substr(data,1,2)
            WHERE data LIKE '__/__/____'
        """)

    # Migração: garante a coluna 'tipo' em bancos criados antes desta feature.
    try:
        conn.execute("ALTER TABLE registros ADD COLUMN tipo TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        conn.rollback()  # coluna já existe

    # Migração: permissão de roteirização por funcionário (admin libera p/ quem quiser).
    try:
        conn.execute("ALTER TABLE funcionarios ADD COLUMN roteirizacao INTEGER DEFAULT 0")
        conn.commit()
    except Exception:
        conn.rollback()  # coluna já existe

    # Migração: intervalos extras (café manhã/tarde e jantar) p/ conformidade com a planilha.
    for _col in ('cafe_manha_inicio', 'cafe_manha_fim', 'cafe_tarde_inicio',
                 'cafe_tarde_fim', 'jantar_inicio', 'jantar_fim'):
        try:
            conn.execute(f"ALTER TABLE registros ADD COLUMN {_col} TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            conn.rollback()  # coluna já existe

    # Seed do cache de geocodificação (uma vez, só se a tabela estiver vazia).
    try:
        if not conn.execute("SELECT 1 FROM clientes_geo LIMIT 1").fetchone():
            import csv as _csv
            seed = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'clientes_geo_seed.csv')
            if os.path.exists(seed):
                with open(seed, encoding='utf-8-sig') as f:
                    for r in _csv.DictReader(f):
                        if not r.get('lat'):
                            continue
                        conn.execute(
                            "INSERT INTO clientes_geo (cod, lat, lng, precisao, atualizado_em) VALUES (?,?,?,?,?)",
                            (str(r['cod']), float(r['lat']), float(r['lng']), r.get('precisao', ''), '')
                        )
                conn.commit()
    except Exception:
        conn.rollback()

    # Migração: feriado deixou de ser marcado pelo funcionário (observacao='FERIADO')
    # e passou a ser um abono administrativo (tipo='feriado').
    conn.execute(
        "UPDATE registros SET tipo='feriado' WHERE observacao='FERIADO' AND (tipo='' OR tipo IS NULL)"
    )

    # Admin: usa env var se definida, senão gera senha aleatória na primeira execução
    senha_admin = os.environ.get('ADMIN_PASSWORD', '')
    existe = conn.execute("SELECT id FROM funcionarios WHERE usuario='admin'").fetchone()
    if not existe:
        if not senha_admin:
            senha_admin = secrets.token_urlsafe(12)
            print("\n" + "="*45)
            print("  ADMIN CRIADO — ANOTE ESTA SENHA:")
            print(f"  Usuário : admin")
            print(f"  Senha   : {senha_admin}")
            print("  (não será exibida novamente)")
            print("="*45 + "\n")
        conn.execute(
            "INSERT INTO funcionarios (nome, usuario, senha_hash, cargo, ativo) VALUES (?,?,?,?,1)",
            ('Administrador', 'admin', _hash_pbkdf2_full(senha_admin), 'ADMIN')
        )
    elif senha_admin:
        # Atualiza hash do admin se env var foi definida (ex: rotação de senha)
        conn.execute(
            "UPDATE funcionarios SET senha_hash=? WHERE usuario='admin'",
            (_hash_pbkdf2_full(senha_admin),)
        )

    conn.commit()
    conn.close()

# ── HASHING ────────────────────────────────────────────
def _hash_pbkdf2_full(senha: str) -> str:
    """Retorna 'pbkdf2:<salt>:<hash>' usando PBKDF2-HMAC-SHA256 com 260k iterações."""
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac('sha256', senha.encode('utf-8'), salt.encode(), 260_000)
    return f"pbkdf2:{salt}:{dk.hex()}"

def verificar_senha(senha: str, stored: str) -> bool:
    """Suporta PBKDF2 (novo) e SHA256 puro (legado), usando compare_digest."""
    if stored.startswith('pbkdf2:'):
        parts = stored.split(':', 2)
        if len(parts) != 3:
            return False
        _, salt, expected = parts
        dk = hashlib.pbkdf2_hmac('sha256', senha.encode('utf-8'), salt.encode(), 260_000)
        return secrets.compare_digest(dk.hex(), expected)
    # Legado: SHA256 sem salt
    return secrets.compare_digest(hashlib.sha256(senha.encode()).hexdigest(), stored)

# ── DATAS ──────────────────────────────────────────────
_RE_ISO = re.compile(r'^\d{4}-\d{2}-\d{2}$')
_RE_BR  = re.compile(r'^(\d{2})/(\d{2})/(\d{4})$')

def normalizar_data(d: str):
    """Aceita YYYY-MM-DD ou DD/MM/YYYY e retorna ISO, ou None se inválida."""
    d = str(d or '').strip()
    if _RE_ISO.match(d):
        return d
    m = _RE_BR.match(d)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None

def _validar_mes(mes: str) -> bool:
    return bool(re.match(r'^\d{4}-(0[1-9]|1[0-2])$', mes))

# ── UTILITÁRIOS ────────────────────────────────────────
def _check_rate_limit(ip: str) -> bool:
    now = time.time()
    attempts = [t for t in _login_attempts.get(ip, []) if now - t < LOGIN_WINDOW]
    if len(attempts) >= LOGIN_MAX:
        _login_attempts[ip] = attempts
        return False
    attempts.append(now)
    _login_attempts[ip] = attempts
    return True

def require_admin_key(f):
    """Decorator: exige header X-Admin-Key quando ADMIN_API_KEY está configurado."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if ADMIN_KEY:
            provided = request.headers.get('X-Admin-Key', '')
            if not secrets.compare_digest(provided, ADMIN_KEY):
                return jsonify({"ok": False, "erro": "Não autorizado"}), 401
        return f(*args, **kwargs)
    return decorated

def _funcionario_do_token():
    """Valida o header Authorization: Bearer <token>. Retorna funcionario_id ou None."""
    auth = request.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return None
    token = auth[7:].strip()
    if not token:
        return None
    conn = get_db()
    row = conn.execute(
        f"""SELECT funcionario_id FROM tokens
           WHERE token=? AND ultimo_uso > {CUTOFF_SQL}""",
        (token,)
    ).fetchone()
    if row:
        conn.execute(f"UPDATE tokens SET ultimo_uso={NOW_SQL} WHERE token=?", (token,))
        conn.commit()
    conn.close()
    return row['funcionario_id'] if row else None

def require_token(f):
    """Decorator: exige token válido e injeta funcionario_id autenticado."""
    @wraps(f)
    def decorated(*args, **kwargs):
        fid = _funcionario_do_token()
        if fid is None:
            return jsonify({"ok": False, "erro": "Sessão expirada. Faça login novamente."}), 401
        return f(fid, *args, **kwargs)
    return decorated

# ── LOGIN ──────────────────────────────────────────────
@app.route('/login', methods=['POST'])
def login():
    ip = request.remote_addr or '0.0.0.0'
    if not _check_rate_limit(ip):
        return jsonify({"ok": False, "erro": "Muitas tentativas. Aguarde 1 minuto."}), 429

    data = request.get_json(silent=True) or {}
    usuario = str(data.get('usuario', '')).strip()[:50]
    senha   = str(data.get('senha', ''))[:100]

    if not usuario or not senha:
        return jsonify({"ok": False, "erro": "Usuário e senha obrigatórios"}), 400

    conn = get_db()
    row = conn.execute(
        "SELECT * FROM funcionarios WHERE usuario=? AND ativo=1", (usuario,)
    ).fetchone()

    if row and verificar_senha(senha, row['senha_hash']):
        token = secrets.token_urlsafe(32)
        conn.execute(
            "INSERT INTO tokens (token, funcionario_id) VALUES (?,?)",
            (token, row['id'])
        )
        # Limpa tokens expirados do funcionário
        conn.execute(
            f"DELETE FROM tokens WHERE funcionario_id=? AND ultimo_uso < {CUTOFF_SQL}",
            (row['id'],)
        )
        conn.commit()
        conn.close()
        return jsonify({
            "ok": True, "id": row["id"], "nome": row["nome"],
            "usuario": row["usuario"], "matricula": row["matricula"],
            "cargo": row["cargo"], "roteirizacao": int(row["roteirizacao"] or 0),
            "token": token
        })
    conn.close()
    return jsonify({"ok": False, "erro": "Usuário ou senha inválidos"}), 401

# ── DADOS ATUAIS DO FUNCIONÁRIO ───────────────────────
@app.route('/me', methods=['GET'])
@require_token
def me(auth_fid):
    """Retorna os dados atuais do funcionário autenticado (ex.: cargo atualizado)."""
    conn = get_db()
    row = conn.execute(
        "SELECT id, nome, usuario, matricula, cargo, roteirizacao FROM funcionarios WHERE id=? AND ativo=1",
        (auth_fid,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({"ok": False, "erro": "Funcionário não encontrado"}), 404
    f = dict(row)
    f["roteirizacao"] = int(f.get("roteirizacao") or 0)
    return jsonify({"ok": True, "funcionario": f})

# ── CLIENTES PARA O MAPA (roteirização) ───────────────
# Diretório (cod, endereço, rota) vem de clientes_enderecos.csv, carregado 1x em
# memória. As coordenadas são geocodificadas SOB DEMANDA (só as rotas usadas) e
# cacheadas na tabela clientes_geo. Cada requisição geocodifica um lote pequeno
# para caber no timeout; o app reabre/atualiza até a rota ficar 100% pronta.
import urllib.request as _urlreq, urllib.parse as _urlparse

_GEO_UA   = "PonttusRoteirizacao/1.0 (gabriellcs71@gmail.com)"
_GEO_MAX  = 6          # geocodificações por requisição (≈ lote dentro do timeout)
_GEO_DELAY = 1.1       # respeita o uso justo do Nominatim (≤ 1 req/s)
_geo_last = [0.0]
_diretorio = {"rows": None}
_ABREV = {"R.": "RUA ", "AV.": "AVENIDA ", "PC.": "PRACA ", "PCA.": "PRACA ",
          "TRV": "TRAVESSA ", "ROD.": "RODOVIA ", "AL.": "ALAMEDA ", "EST.": "ESTRADA "}

def _carregar_diretorio():
    """Lê clientes_enderecos.csv (cod -> dados do cliente) uma vez."""
    if _diretorio["rows"] is not None:
        return _diretorio["rows"]
    import csv as _csv
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, 'clientes_enderecos.csv')
    rows = {}
    if os.path.exists(path):
        with open(path, encoding='utf-8-sig') as f:
            for r in _csv.DictReader(f):
                cod = str(r.get('cod', '')).strip()
                if not cod:
                    continue
                rows[cod] = {
                    'cod': cod,
                    'nome': r.get('fantasia') or r.get('razao') or '',
                    'endereco': r.get('endereco', ''),
                    'numero': r.get('numero', ''),
                    'bairro': r.get('bairro', ''),
                    'cidade': r.get('cidade', ''),
                    'uf': r.get('uf', '') or 'MG',
                    'vendedores': [v for v in (r.get('vendedores', '') or '').split(';') if v],
                }
    _diretorio["rows"] = rows
    return rows

def _nominatim(params):
    dt = time.time() - _geo_last[0]
    if dt < _GEO_DELAY:
        time.sleep(_GEO_DELAY - dt)
    url = "https://nominatim.openstreetmap.org/search?" + _urlparse.urlencode(params)
    req = _urlreq.Request(url, headers={"User-Agent": _GEO_UA})
    _geo_last[0] = time.time()
    try:
        d = json.load(_urlreq.urlopen(req, timeout=10))
        return (float(d[0]["lat"]), float(d[0]["lon"])) if d else None
    except Exception:
        return None

def _geocodificar(rec):
    """Retorna (lat, lng, precisao). Tenta rua+nº, depois rua, depois centro da cidade."""
    rua = rec['endereco'].upper()
    for k, v in _ABREV.items():
        if rua.startswith(k):
            rua = v + rua[len(k):]
    rua = re.sub(r"\s+", " ", rua).strip()
    cidade, uf, num = rec['cidade'], rec['uf'], rec['numero']
    r = _nominatim({"street": (num + " " + rua).strip(), "city": cidade, "state": uf,
                    "country": "Brasil", "format": "json", "limit": 1, "countrycodes": "br"})
    if r:
        return r[0], r[1], "rua"
    r = _nominatim({"street": rua, "city": cidade, "state": uf, "country": "Brasil",
                    "format": "json", "limit": 1, "countrycodes": "br"})
    if r:
        return r[0], r[1], "rua"
    r = _nominatim({"city": cidade, "state": uf, "country": "Brasil",
                    "format": "json", "limit": 1, "countrycodes": "br"})
    if r:
        return r[0], r[1], "cidade"
    return None, None, "nenhuma"

@app.route('/clientes_mapa', methods=['GET'])
@require_token
def clientes_mapa(auth_fid):
    """Pontos da rota para o mapa. Geocodifica sob demanda e cacheia.
    ?vend=<codigo> (default: matrícula do logado) · ?cidade=<nome> filtro opcional.
    Resposta inclui 'preparando' (bool) e 'prontos'/'total' p/ o app atualizar até concluir."""
    conn = get_db()
    me = conn.execute(
        "SELECT usuario, matricula, roteirizacao FROM funcionarios WHERE id=?", (auth_fid,)
    ).fetchone()
    autorizado = me and ((me['usuario'] or '').strip().lower() == 'admin'
                         or int(me['roteirizacao'] or 0) == 1)
    if not autorizado:
        conn.close()
        return jsonify({"ok": False, "erro": "Acesso restrito."}), 403

    vend = (request.args.get('vend') or '').strip()
    cidade = (request.args.get('cidade') or '').strip().upper()
    if not vend:
        vend = str(me['matricula']).strip() if me['matricula'] else ''

    diret = _carregar_diretorio()
    todos = vend.lower() == 'todos'
    rota = [rec for rec in diret.values()
            if (todos or (vend and vend in rec['vendedores']))
            and (not cidade or rec['cidade'].upper() == cidade)]

    if not rota:
        conn.close()
        return jsonify({"ok": True, "vend": vend, "total": 0, "prontos": 0,
                        "preparando": False, "clientes": []})

    # Coordenadas já em cache para os clientes desta rota.
    cods = [r['cod'] for r in rota]
    cache = {}
    CHUNK = 400
    for i in range(0, len(cods), CHUNK):
        parte = cods[i:i + CHUNK]
        ph = ",".join("?" * len(parte))
        for row in conn.execute(
            f"SELECT cod, lat, lng, precisao FROM clientes_geo WHERE cod IN ({ph})", parte
        ).fetchall():
            cache[str(row['cod'])] = row

    pendentes = [r for r in rota if r['cod'] not in cache]

    # Geocodifica um lote pequeno agora (cabe no timeout); o resto vem nas próximas chamadas.
    for rec in pendentes[:_GEO_MAX]:
        lat, lng, prec = _geocodificar(rec)
        try:
            conn.execute(
                "INSERT INTO clientes_geo (cod, lat, lng, precisao, atualizado_em) VALUES (?,?,?,?,?)",
                (rec['cod'], lat, lng, prec, time.strftime('%Y-%m-%d %H:%M:%S'))
            )
            conn.commit()
        except INTEGRITY_ERRORS:
            conn.rollback()
        cache[rec['cod']] = {'cod': rec['cod'], 'lat': lat, 'lng': lng, 'precisao': prec}

    conn.close()

    # Monta a saída: só clientes que já têm coordenada.
    clientes = []
    for rec in rota:
        c = cache.get(rec['cod'])
        if not c or c['lat'] is None:
            continue
        clientes.append({**rec, 'lat': float(c['lat']), 'lng': float(c['lng']),
                         'precisao': c['precisao']})

    faltam = len(pendentes) - min(len(pendentes), _GEO_MAX)
    return jsonify({
        "ok": True, "vend": vend, "total": len(rota), "prontos": len(clientes),
        "preparando": faltam > 0, "clientes": clientes,
    })

# ── TROCAR SENHA (funcionário) ────────────────────────
@app.route('/senha', methods=['POST'])
@require_token
def trocar_senha(auth_fid):
    """O funcionário autenticado troca a própria senha (exige a senha atual)."""
    data = request.get_json(silent=True) or {}
    atual = str(data.get('senha_atual', '') or '')[:100]
    nova  = str(data.get('nova_senha', '') or '')[:100]
    if len(nova) < 6:
        return jsonify({"ok": False, "erro": "A nova senha deve ter no mínimo 6 caracteres"}), 400

    conn = get_db()
    row = conn.execute("SELECT senha_hash FROM funcionarios WHERE id=?", (auth_fid,)).fetchone()
    if not row or not verificar_senha(atual, row['senha_hash']):
        conn.close()
        return jsonify({"ok": False, "erro": "Senha atual incorreta"}), 401
    conn.execute("UPDATE funcionarios SET senha_hash=? WHERE id=?",
                 (_hash_pbkdf2_full(nova), auth_fid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# ── REGISTROS ─────────────────────────────────────────
@app.route('/registros', methods=['POST'])
@require_token
def salvar_registros(auth_fid):
    data = request.get_json(silent=True) or {}
    registros = data.get('registros', [])

    if not isinstance(registros, list) or not registros or len(registros) > 500:
        return jsonify({"ok": False, "erro": "Dados inválidos"}), 400

    funcionario_id = auth_fid  # ignora id do payload: usa sempre o do token

    conn = get_db()
    salvos = 0
    for reg in registros:
        data_iso = normalizar_data(reg.get('data', ''))
        if not data_iso:
            continue

        campos = (
            str(reg.get('cidade')    or '')[:100],
            str(reg.get('entrada')   or '')[:5],
            str(reg.get('cafe_manha_inicio') or '')[:5],
            str(reg.get('cafe_manha_fim')    or '')[:5],
            str(reg.get('almoco_inicio') or '')[:5],
            str(reg.get('almoco_fim')    or '')[:5],
            str(reg.get('cafe_tarde_inicio') or '')[:5],
            str(reg.get('cafe_tarde_fim')    or '')[:5],
            str(reg.get('jantar_inicio') or '')[:5],
            str(reg.get('jantar_fim')    or '')[:5],
            str(reg.get('saida')     or '')[:5],
            str(reg.get('observacao')or '')[:500],
        )

        # Funcionário só grava dia de trabalho (tipo=''); se o dia já é um abono
        # lançado pelo admin, o WHERE impede a sobrescrita — o abono prevalece.
        conn.execute(f"""
            INSERT INTO registros (funcionario_id, data, cidade, entrada,
                cafe_manha_inicio, cafe_manha_fim, almoco_inicio, almoco_fim,
                cafe_tarde_inicio, cafe_tarde_fim, jantar_inicio, jantar_fim,
                saida, observacao, tipo)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,'')
            ON CONFLICT(funcionario_id, data) DO UPDATE SET
                cidade=excluded.cidade, entrada=excluded.entrada,
                cafe_manha_inicio=excluded.cafe_manha_inicio, cafe_manha_fim=excluded.cafe_manha_fim,
                almoco_inicio=excluded.almoco_inicio, almoco_fim=excluded.almoco_fim,
                cafe_tarde_inicio=excluded.cafe_tarde_inicio, cafe_tarde_fim=excluded.cafe_tarde_fim,
                jantar_inicio=excluded.jantar_inicio, jantar_fim=excluded.jantar_fim,
                saida=excluded.saida, observacao=excluded.observacao,
                enviado_em={NOW_SQL}
            WHERE registros.tipo='' OR registros.tipo IS NULL
        """, (funcionario_id, data_iso, *campos))
        salvos += 1

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "salvos": salvos})

@app.route('/registros', methods=['GET'])
@require_token
def get_registros(auth_fid):
    mes = request.args.get('mes', '')
    conn = get_db()
    if mes:
        if not _validar_mes(mes):
            conn.close()
            return jsonify({"ok": False, "erro": "Formato de mês inválido (use YYYY-MM)"}), 400
        rows = conn.execute(
            "SELECT * FROM registros WHERE funcionario_id=? AND data LIKE ? ORDER BY data",
            (auth_fid, f"{mes}%")
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM registros WHERE funcionario_id=? ORDER BY data DESC LIMIT 120",
            (auth_fid,)
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ── CHECKLIST DO VEÍCULO (somente motoristas) ─────────
def _eh_motorista(fid, conn):
    row = conn.execute("SELECT cargo FROM funcionarios WHERE id=?", (fid,)).fetchone()
    return bool(row) and (row['cargo'] or '').strip().upper() == 'MOTORISTA'

@app.route('/checklist', methods=['POST'])
@require_token
def salvar_checklist(auth_fid):
    data = request.get_json(silent=True) or {}
    data_iso = normalizar_data(data.get('data', ''))
    if not data_iso:
        return jsonify({"ok": False, "erro": "Data inválida"}), 400
    conn = get_db()
    if not _eh_motorista(auth_fid, conn):
        conn.close()
        return jsonify({"ok": False, "erro": "Apenas motoristas podem enviar o checklist."}), 403
    veiculo = str(data.get('veiculo', '') or '')[:60]
    km      = str(data.get('km', '') or '')[:20]
    prox    = str(data.get('prox_troca_oleo', '') or '')[:20]
    itens   = data.get('itens', [])
    if not isinstance(itens, list) or len(itens) > 60:
        conn.close()
        return jsonify({"ok": False, "erro": "Itens inválidos"}), 400
    limpos = []
    for it in itens:
        if not isinstance(it, dict):
            continue
        limpos.append({
            'item': str(it.get('item', ''))[:80],
            'status': 'problema' if str(it.get('status', '')) == 'problema' else 'ok',
            'obs': str(it.get('obs', '') or '')[:300],
        })
    itens_json = json.dumps(limpos, ensure_ascii=False)
    conn.execute(f"""
        INSERT INTO checklists (funcionario_id, data, veiculo, km, prox_troca_oleo, itens)
        VALUES (?,?,?,?,?,?)
        ON CONFLICT(funcionario_id, data) DO UPDATE SET
            veiculo=excluded.veiculo, km=excluded.km, prox_troca_oleo=excluded.prox_troca_oleo,
            itens=excluded.itens, criado_em={NOW_SQL}
    """, (auth_fid, data_iso, veiculo, km, prox, itens_json))

    # Gera as manutenções a partir dos itens marcados como "problema".
    # Preserva o status de pendências já resolvidas; remove as que não são mais problema.
    problemas = [it for it in limpos if it['status'] == 'problema']
    nomes_prob = set(p['item'] for p in problemas)
    pend = conn.execute(
        "SELECT id, item FROM manutencoes WHERE funcionario_id=? AND data=? AND status='pendente'",
        (auth_fid, data_iso)
    ).fetchall()
    for row in pend:
        if row['item'] not in nomes_prob:
            conn.execute("DELETE FROM manutencoes WHERE id=?", (row['id'],))
    for p in problemas:
        conn.execute("""
            INSERT INTO manutencoes (funcionario_id, data, veiculo, item, obs)
            VALUES (?,?,?,?,?)
            ON CONFLICT(funcionario_id, data, item) DO UPDATE SET
                veiculo=excluded.veiculo, obs=excluded.obs
        """, (auth_fid, data_iso, veiculo, p['item'], p['obs']))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route('/checklist', methods=['GET'])
@require_token
def get_checklist(auth_fid):
    data_iso = normalizar_data(request.args.get('data', ''))
    conn = get_db()
    if not _eh_motorista(auth_fid, conn):
        conn.close()
        return jsonify({"ok": False, "erro": "Apenas motoristas"}), 403
    row = None
    if data_iso:
        row = conn.execute("SELECT * FROM checklists WHERE funcionario_id=? AND data=?",
                           (auth_fid, data_iso)).fetchone()
    conn.close()
    if not row:
        return jsonify({"ok": True, "checklist": None})
    d = dict(row)
    try:
        d['itens'] = json.loads(d.get('itens') or '[]')
    except Exception:
        d['itens'] = []
    return jsonify({"ok": True, "checklist": d})

@app.route('/admin/checklists', methods=['GET'])
@require_admin_key
def admin_checklists():
    mes = request.args.get('mes', '')
    conn = get_db()
    query = """
        SELECT c.*, f.nome as funcionario_nome, f.matricula as funcionario_matricula
        FROM checklists c JOIN funcionarios f ON f.id = c.funcionario_id
    """
    params = []
    if mes:
        if not _validar_mes(mes):
            conn.close()
            return jsonify({"ok": False, "erro": "Formato de mês inválido (use YYYY-MM)"}), 400
        query += " WHERE c.data LIKE ?"
        params.append(f"{mes}%")
    query += " ORDER BY c.data DESC, f.nome"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d['itens'] = json.loads(d.get('itens') or '[]')
        except Exception:
            d['itens'] = []
        out.append(d)
    return jsonify(out)

@app.route('/admin/manutencoes', methods=['GET'])
@require_admin_key
def admin_manutencoes():
    """Lista as manutenções (problemas reportados). status=pendente|resolvido|todos.
    Ordenadas por horário do reporte (mais antigas primeiro)."""
    status = request.args.get('status', 'pendente')
    conn = get_db()
    query = """
        SELECT m.*, f.nome as funcionario_nome, f.matricula as funcionario_matricula
        FROM manutencoes m JOIN funcionarios f ON f.id = m.funcionario_id
    """
    params = []
    if status in ('pendente', 'resolvido'):
        query += " WHERE m.status=?"
        params.append(status)
    query += " ORDER BY m.criado_em ASC, m.data ASC"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/admin/manutencoes/resolver', methods=['POST'])
@require_admin_key
def admin_manutencao_resolver():
    data = request.get_json(silent=True) or {}
    try:
        mid = int(data.get('id'))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "erro": "id inválido"}), 400
    novo = 'pendente' if str(data.get('acao', '')) == 'reabrir' else 'resolvido'
    conn = get_db()
    quando = NOW_SQL if novo == 'resolvido' else "''"
    conn.execute(f"UPDATE manutencoes SET status=?, resolvido_em={quando} WHERE id=?", (novo, mid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "status": novo})

# ── ADMIN: FUNCIONÁRIOS ───────────────────────────────
@app.route('/admin/funcionarios', methods=['GET'])
@require_admin_key
def listar_funcionarios():
    conn = get_db()
    try:
        limit  = min(int(request.args.get('limit', 100)), 500)
        offset = max(int(request.args.get('offset', 0)), 0)
    except ValueError:
        conn.close()
        return jsonify({"ok": False, "erro": "Parâmetros inválidos"}), 400
    rows = conn.execute(
        "SELECT id, nome, usuario, matricula, cargo, roteirizacao, ativo, criado_em FROM funcionarios ORDER BY nome LIMIT ? OFFSET ?",
        (limit, offset)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/admin/funcionarios', methods=['POST'])
@require_admin_key
def criar_funcionario():
    data    = request.get_json(silent=True) or {}
    nome    = str(data.get('nome',    '') or '').strip()[:100]
    usuario = str(data.get('usuario', '') or '').strip().lower()[:50]
    senha   = str(data.get('senha',   '') or '')[:100]

    if not nome or not usuario or not senha:
        return jsonify({"ok": False, "erro": "Nome, usuário e senha são obrigatórios"}), 400
    if not re.match(r'^[a-z0-9._-]{3,50}$', usuario):
        return jsonify({"ok": False, "erro": "Usuário deve ter 3-50 caracteres minúsculos alfanuméricos (. _ - permitidos)"}), 400
    if len(senha) < 6:
        return jsonify({"ok": False, "erro": "Senha deve ter no mínimo 6 caracteres"}), 400

    matricula = str(data.get('matricula', '') or '')[:20]
    cargo     = str(data.get('cargo',     '') or '')[:50]
    roteiriz  = int(bool(data.get('roteirizacao', 0)))

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO funcionarios (nome, usuario, senha_hash, matricula, cargo, roteirizacao) VALUES (?,?,?,?,?,?)",
            (nome, usuario, _hash_pbkdf2_full(senha), matricula, cargo, roteiriz)
        )
        conn.commit()
        return jsonify({"ok": True})
    except INTEGRITY_ERRORS:
        conn.rollback()
        return jsonify({"ok": False, "erro": "Usuário já existe"}), 409
    finally:
        conn.close()

@app.route('/admin/funcionarios/<int:fid>', methods=['PUT'])
@require_admin_key
def atualizar_funcionario(fid):
    data    = request.get_json(silent=True) or {}
    nome    = str(data.get('nome',    '') or '').strip()[:100]
    usuario = str(data.get('usuario', '') or '').strip().lower()[:50]

    if not nome or not usuario:
        return jsonify({"ok": False, "erro": "Nome e usuário são obrigatórios"}), 400

    matricula = str(data.get('matricula', '') or '')[:20]
    cargo     = str(data.get('cargo',     '') or '')[:50]
    ativo     = int(bool(data.get('ativo', 1)))
    roteiriz  = int(bool(data.get('roteirizacao', 0)))

    conn = get_db()
    if data.get('senha'):
        senha = str(data['senha'])[:100]
        if len(senha) < 6:
            conn.close()
            return jsonify({"ok": False, "erro": "Senha deve ter no mínimo 6 caracteres"}), 400
        conn.execute(
            "UPDATE funcionarios SET nome=?, usuario=?, senha_hash=?, matricula=?, cargo=?, roteirizacao=?, ativo=? WHERE id=?",
            (nome, usuario, _hash_pbkdf2_full(senha), matricula, cargo, roteiriz, ativo, fid)
        )
    else:
        conn.execute(
            "UPDATE funcionarios SET nome=?, usuario=?, matricula=?, cargo=?, roteirizacao=?, ativo=? WHERE id=?",
            (nome, usuario, matricula, cargo, roteiriz, ativo, fid)
        )
    if not ativo:
        conn.execute("DELETE FROM tokens WHERE funcionario_id=?", (fid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route('/admin/registros', methods=['GET'])
@require_admin_key
def admin_registros():
    mes = request.args.get('mes', '')
    conn = get_db()
    query = """
        SELECT r.*, f.nome as funcionario_nome, f.matricula as funcionario_matricula
        FROM registros r
        JOIN funcionarios f ON f.id = r.funcionario_id
    """
    params = []
    if mes:
        if not _validar_mes(mes):
            conn.close()
            return jsonify({"ok": False, "erro": "Formato de mês inválido (use YYYY-MM)"}), 400
        query += " WHERE r.data LIKE ?"
        params.append(f"{mes}%")
    query += " ORDER BY f.nome, r.data"
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/admin/exportar', methods=['GET'])
@require_admin_key
def admin_exportar():
    """Gera a planilha de ponto no modelo da empresa: um bloco por funcionário,
    com os dias do mês preenchidos a partir dos registros. Param: mes=YYYY-MM."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    mes = request.args.get('mes', '')
    if not _validar_mes(mes):
        return jsonify({"ok": False, "erro": "Use mes=YYYY-MM"}), 400
    ano, m = int(mes[:4]), int(mes[5:7])
    ndias = calendar.monthrange(ano, m)[1]

    ABONO_LABEL = {'falta': 'FALTA', 'atestado': 'ATESTADO', 'folga': 'FOLGA', 'feriado': 'FERIADO'}
    DIAS = ['SEGUNDA', 'TERÇA', 'QUARTA', 'QUINTA', 'SEXTA', 'SÁBADO', 'DOMINGO']
    MESNOME = ['', 'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
               'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']

    conn = get_db()
    funcs = conn.execute(
        "SELECT id, nome, matricula, cargo FROM funcionarios "
        "WHERE ativo=1 AND usuario <> 'admin' ORDER BY nome"
    ).fetchall()
    regs = conn.execute(
        "SELECT * FROM registros WHERE data LIKE ? ORDER BY data", (f"{mes}%",)
    ).fetchall()
    conn.close()

    # registros por funcionário e dia
    por_func = {}
    for r in regs:
        d = dict(r)
        por_func.setdefault(d['funcionario_id'], {})[d['data']] = d

    def to_time(s):
        s = (s or '').strip()
        if not s or ':' not in s:
            return None
        try:
            hh, mm = s.split(':')[:2]
            return _dt.time(int(hh), int(mm))
        except Exception:
            return None

    wb = Workbook()
    ws = wb.active
    ws.title = f"{MESNOME[m]} {ano}"

    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='BBBBBB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill('solid', fgColor='1F2937')
    head_font = Font(bold=True, color='FFFFFF')

    widths = {'A': 11, 'B': 11, 'C': 20, 'D': 8, 'E': 8, 'F': 8, 'G': 8,
              'H': 8, 'I': 8, 'J': 8, 'K': 8, 'L': 8, 'M': 8, 'N': 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    top = 1
    for f in funcs:
        fid, nome, matric, cargo = f['id'], f['nome'], f['matricula'], (f['cargo'] or '')
        dados = por_func.get(fid, {})

        ws.cell(top, 3, 'CONTROLE DE PONTO / TRABALHO EXTERNO').font = bold
        ws.cell(top + 2, 3, nome.upper()).font = bold
        # cabeçalhos
        ws.cell(top + 3, 1, 'DATA'); ws.cell(top + 3, 3, 'LOCAL DA ENTRADA')
        ws.cell(top + 3, 4, 'HORÁRIOS DE TRABALHO'); ws.cell(top + 3, 14, 'DIA')
        sub1 = top + 4
        ws.cell(sub1, 4, 'ENTRADA'); ws.cell(sub1, 5, 'DESCANSO / CAFÉ')
        ws.cell(sub1, 7, 'ALMOÇO'); ws.cell(sub1, 9, 'DESCANSO / CAFÉ')
        ws.cell(sub1, 11, 'JANTAR'); ws.cell(sub1, 13, 'SAÍDA'); ws.cell(sub1, 14, '8:48:00')
        sub2 = top + 5
        for c, t in [(5, 'INÍCIO'), (6, 'FINAL'), (7, 'INÍCIO'), (8, 'FINAL'),
                     (9, 'INÍCIO'), (10, 'FINAL'), (11, 'INÍCIO'), (12, 'FINAL')]:
            ws.cell(sub2, c, t)
        for rr in (top + 3, sub1, sub2):
            for c in range(1, 15):
                cell = ws.cell(rr, c)
                cell.font = head_font; cell.fill = head_fill
                cell.alignment = center; cell.border = border

        first_data = top + 6
        for i in range(ndias):
            dia = i + 1
            row = first_data + i
            data_iso = f"{ano:04d}-{m:02d}-{dia:02d}"
            dt = _dt.date(ano, m, dia)
            reg = dados.get(data_iso, {})
            tipo = (reg.get('tipo') or '').strip().lower()

            ws.cell(row, 1, DIAS[dt.weekday()])
            ws.cell(row, 2, dt).number_format = 'DD/MM/YYYY'
            if tipo in ABONO_LABEL:
                ws.cell(row, 3, ABONO_LABEL[tipo]).font = bold
            else:
                ws.cell(row, 3, reg.get('cidade') or '')
            colmap = {4: 'entrada', 5: 'cafe_manha_inicio', 6: 'cafe_manha_fim',
                      7: 'almoco_inicio', 8: 'almoco_fim', 9: 'cafe_tarde_inicio',
                      10: 'cafe_tarde_fim', 11: 'jantar_inicio', 12: 'jantar_fim',
                      13: 'saida'}
            for c, campo in colmap.items():
                t = to_time(reg.get(campo))
                if t is not None:
                    cell = ws.cell(row, c, t)
                    cell.number_format = 'HH:MM'
            # total do dia (mesma lógica da planilha original)
            ws.cell(row, 14,
                f'=IF(C{row}="FOLGA",-TIME(8,48,0),IF(M{row}=0,0,IF(C{row}&D{row}="",0,'
                f'IF(OR(C{row}="FÉRIAS",C{row}="FALTA",C{row}="ATESTADO",C{row}="FERIADO"),0,'
                f'(G{row}-D{row})+(K{row}-H{row})+(M{row}-L{row})-TIME(8,48,0))))))'
            ).number_format = '[h]:mm'
            for c in range(1, 15):
                ws.cell(row, c).border = border

        last_data = first_data + ndias - 1
        foot = last_data + 1
        ws.cell(foot, 11, 'TOTAL DO MÊS').font = bold
        tot = ws.cell(foot, 14, f'=SUM(N{first_data}:N{last_data})')
        tot.number_format = '[h]:mm'; tot.font = bold

        ws.cell(foot + 2, 1, 'NOME COMPLETO:').font = bold
        ws.cell(foot + 2, 3, f'{nome.upper()}   {matric or ""}')
        ws.cell(foot + 3, 1, 'FUNÇÃO:').font = bold
        ws.cell(foot + 3, 3, cargo.upper())

        top += 42  # próximo bloco

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"CONTROLE_PONTO_{MESNOME[m]}_{ano}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/abono', methods=['POST'])
@require_admin_key
def admin_abono():
    """Lança ou remove um abono (falta/atestado/folga/feriado) num dia.
    Somente o admin. O abono prevalece: zera as horas trabalhadas do dia.
    Enviar tipo='' remove o abono (o dia volta a ficar vazio)."""
    data = request.get_json(silent=True) or {}
    try:
        fid = int(data.get('funcionario_id'))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "erro": "funcionario_id inválido"}), 400

    data_iso = normalizar_data(data.get('data', ''))
    if not data_iso:
        return jsonify({"ok": False, "erro": "Data inválida (use YYYY-MM-DD)"}), 400

    tipo = str(data.get('tipo', '') or '').strip().lower()
    if tipo and tipo not in ABONO_TIPOS:
        return jsonify({"ok": False, "erro": "Tipo inválido"}), 400

    conn = get_db()
    func = conn.execute("SELECT id FROM funcionarios WHERE id=?", (fid,)).fetchone()
    if not func:
        conn.close()
        return jsonify({"ok": False, "erro": "Funcionário não encontrado"}), 404

    if tipo == '':
        # Remove o abono: apaga o registro do dia (volta a ficar vazio).
        conn.execute("DELETE FROM registros WHERE funcionario_id=? AND data=?", (fid, data_iso))
    else:
        # Marca o abono zerando horas/cidade — o abono prevalece sobre o trabalho.
        conn.execute(f"""
            INSERT INTO registros (funcionario_id, data, cidade, entrada,
                almoco_inicio, almoco_fim, saida, observacao, tipo)
            VALUES (?,?,'','','','','','',?)
            ON CONFLICT(funcionario_id, data) DO UPDATE SET
                tipo=excluded.tipo, cidade='', entrada='', almoco_inicio='',
                almoco_fim='', saida='', enviado_em={NOW_SQL}
        """, (fid, data_iso, tipo))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "tipo": tipo})

# ── PAINEL ADMIN (WEB) ────────────────────────────────
@app.route('/')
def raiz():
    return render_template('painel.html')

@app.route('/painel')
def painel():
    return render_template('painel.html')

@app.route('/ping', methods=['GET'])
def ping():
    return jsonify({"ok": True, "msg": "Ponttus servidor online", "db": "postgres" if USE_PG else "sqlite"})

# Garante schema também sob gunicorn (Render), não só em execução direta
init_db()

if __name__ == '__main__':
    port  = int(os.environ.get('PORT', 5000))
    host  = os.environ.get('HOST', '0.0.0.0')
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(host=host, port=port, debug=debug)
